from django.shortcuts import render
from rest_framework import viewsets, generics, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db.models import Sum
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Wallet, PortExpense, DigitalVoucher, WalletTopUp, TallyLog
from .serializers import (
    WalletSerializer, PortExpenseSerializer, DigitalVoucherSerializer,
    WalletTopUpSerializer, TallyLogSerializer
)
from authentication.permissions import (
    IsSupervisorOrAbove, CanApproveFinancial, CanManageWallets,
    IsManagerOrAdmin, IsAccountantOrAdmin, HasWallet
)

class PortExpenseViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing port expenses
    """
    queryset = PortExpense.objects.all()
    serializer_class = PortExpenseSerializer
    permission_classes = [IsSupervisorOrAbove]
    
    def get_queryset(self):
        queryset = PortExpense.objects.all()
        user = self.request.user
        status_filter = self.request.query_params.get('status', None)
        
        # Non-admin users can only see their own expenses
        if hasattr(user, 'role') and user.role != 'admin':
            queryset = queryset.filter(user=user)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset

class DigitalVoucherViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing digital vouchers
    """
    queryset = DigitalVoucher.objects.all()
    serializer_class = DigitalVoucherSerializer
    permission_classes = [IsSupervisorOrAbove]
    
    def get_queryset(self):
        queryset = DigitalVoucher.objects.all()
        user = self.request.user
        status_filter = self.request.query_params.get('status', None)
        
        # Non-admin users can only see their own vouchers
        if hasattr(user, 'role') and user.role != 'admin':
            queryset = queryset.filter(user=user)
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset

class WalletTopUpViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing wallet top-ups
    """
    queryset = WalletTopUp.objects.all()
    serializer_class = WalletTopUpSerializer
    permission_classes = [CanManageWallets]

class TallyLogViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Tally logs
    """
    queryset = TallyLog.objects.all()
    serializer_class = TallyLogSerializer
    permission_classes = [IsAccountantOrAdmin]
    
    def get_queryset(self):
        queryset = TallyLog.objects.all()
        entry_type = self.request.query_params.get('entry_type', None)
        
        if entry_type:
            queryset = queryset.filter(entry_type=entry_type)
            
        return queryset

class WalletBalanceView(generics.GenericAPIView):
    """
    View to get current wallet balance for authenticated user
    """
    permission_classes = [HasWallet]
    
    def get(self, request):
        user = request.user
        
        # Ensure user has wallet access
        if user.role == 'accountant':
            return Response({'error': 'Accountants do not have wallets'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        balance = Wallet.get_balance(user)
        
        return Response({
            'user_id': user.id,
            'username': user.username,
            'balance': balance,
            'last_updated': timezone.now()
        })

class WalletTransactionsView(generics.ListAPIView):
    """
    View to get wallet transactions for authenticated user
    """
    serializer_class = WalletSerializer
    permission_classes = [HasWallet]
    
    def get_queryset(self):
        user = self.request.user
        
        # Ensure user has wallet access
        if user.role == 'accountant':
            return Wallet.objects.none()
        
        # All wallet holders (admin, manager, supervisor) should only see their own transactions
        return Wallet.objects.filter(user=user)

class ApproveExpenseView(generics.UpdateAPIView):
    """
    View to approve/reject port expenses
    """
    queryset = PortExpense.objects.all()
    serializer_class = PortExpenseSerializer
    permission_classes = []  # Allow all authenticated users, check roles in the method
    
    def patch(self, request, *args, **kwargs):
        # Check authentication
        if not request.user or not request.user.is_authenticated:
            return Response(
                {'error': 'Authentication required'}, 
                status=status.HTTP_401_UNAUTHORIZED
            )
            
        expense = self.get_object()
        user = request.user
        action = request.data.get('action')  # 'approve' or 'reject'
        comments = request.data.get('comments', '')
        
        if action not in ['approve', 'reject', 'finalize']:
            return Response(
                {'error': 'Action must be approve, reject, or finalize'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if action == 'approve' and (user.is_manager or user.is_admin):
            expense.status = 'approved'
            expense.reviewed_by = user
            expense.review_comments = comments
            
        elif action == 'finalize' and user.is_accountant:
            if expense.status != 'approved':
                return Response(
                    {'error': 'Expense must be approved before finalizing'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            expense.status = 'finalized'
            expense.approved_by = user
            
            # Create wallet debit transaction when expense is finalized
            from .models import Wallet, TallyLog
            existing_transaction = Wallet.objects.filter(
                user=expense.user,
                reference='expense',
                reference_id=str(expense.id)
            ).first()
            
            if not existing_transaction:
                Wallet.objects.create(
                    user=expense.user,
                    action='debit',
                    amount=expense.total_amount,
                    reference='expense',
                    reference_id=str(expense.id),
                    approved_by=user,
                    description=f"Port expense - {expense.description}"
                )
                
                # Create Tally log entry for finalized expense
                TallyLog.objects.create(
                    entry_type='expense',
                    reference_id=str(expense.id),
                    tally_voucher_number=f"PE{expense.id:06d}",  # Generate PE000001 format
                    amount=expense.total_amount,
                    description=f"Port expense - {expense.description}",
                    logged_by=user
                )
            
        elif action == 'reject':
            expense.status = 'rejected'
            expense.reviewed_by = user
            expense.review_comments = comments
            
        else:
            return Response(
                {'error': 'Insufficient permissions for this action'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        expense.save()
        serializer = self.get_serializer(expense)
        return Response(serializer.data)

class ApproveVoucherView(generics.UpdateAPIView):
    """
    View to approve/decline digital vouchers
    """
    queryset = DigitalVoucher.objects.all()
    serializer_class = DigitalVoucherSerializer
    permission_classes = []  # Allow all authenticated users, check roles in the method
    
    def patch(self, request, *args, **kwargs):
        # Check authentication
        if not request.user or not request.user.is_authenticated:
            return Response(
                {'error': 'Authentication required'}, 
                status=status.HTTP_401_UNAUTHORIZED
            )
            
        voucher = self.get_object()
        user = request.user
        action = request.data.get('action')  # 'approve', 'decline', or 'log'
        comments = request.data.get('comments', '')
        tally_reference = request.data.get('tally_reference', '')
        
        if action not in ['approve', 'decline', 'log']:
            return Response(
                {'error': 'Action must be approve, decline, or log'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if action == 'approve' and user.is_admin:
            voucher.status = 'approved'
            voucher.approved_by = user
            voucher.approval_comments = comments
            
        elif action == 'log' and user.is_accountant:
            if voucher.status != 'approved':
                return Response(
                    {'error': 'Voucher must be approved before logging'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            voucher.status = 'logged'
            voucher.logged_by = user
            voucher.tally_reference = tally_reference
            
            # Create Tally log entry
            TallyLog.objects.create(
                entry_type='voucher',
                reference_id=str(voucher.id),
                tally_voucher_number=tally_reference,
                amount=voucher.amount,
                description=f"Digital voucher - {voucher.expense_category}",
                logged_by=user
            )
            
        elif action == 'decline':
            voucher.status = 'declined'
            voucher.approved_by = user
            voucher.approval_comments = comments
            
        else:
            return Response(
                {'error': 'Insufficient permissions for this action'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        voucher.save()
        serializer = self.get_serializer(voucher)
        return Response(serializer.data)

class PortExpenseExcelExportView(generics.GenericAPIView):
    """
    Export Port Expenses to Excel - Weekly Report for Accountants
    """
    permission_classes = [IsAccountantOrAdmin]
    
    def get(self, request):
        # Get week parameter or default to current week
        week_str = request.query_params.get('week')
        if week_str:
            try:
                start_date = datetime.strptime(week_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        else:
            # Default to current week (Monday to Sunday)
            today = timezone.now().date()
            start_date = today - timedelta(days=today.weekday())
        
        end_date = start_date + timedelta(days=6)
        
        # Get finalized port expenses for the week
        expenses = PortExpense.objects.filter(
            status='finalized',
            created_at__date__range=[start_date, end_date]
        ).order_by('created_at')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Port Expenses"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            'Date', 'User', 'Vehicle', 'Vehicle Number', 'Gate No', 
            'CISF Amount', 'KPT Amount', 'Customs Amount', 'Road Tax Days',
            'Road Tax Amount', 'Other Charges', 'Total Amount', 'Description',
            'Approved By', 'Created At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, expense in enumerate(expenses, 2):
            ws.cell(row=row_num, column=1, value=expense.date_time.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=2, value=expense.user.get_full_name())
            ws.cell(row=row_num, column=3, value=expense.vehicle)
            ws.cell(row=row_num, column=4, value=expense.vehicle_number)
            ws.cell(row=row_num, column=5, value=expense.get_gate_no_display())
            ws.cell(row=row_num, column=6, value=float(expense.cisf_amount))
            ws.cell(row=row_num, column=7, value=float(expense.kpt_amount))
            ws.cell(row=row_num, column=8, value=float(expense.customs_amount))
            ws.cell(row=row_num, column=9, value=expense.road_tax_days)
            ws.cell(row=row_num, column=10, value=float(expense.road_tax_amount))
            ws.cell(row=row_num, column=11, value=float(expense.other_charges))
            ws.cell(row=row_num, column=12, value=float(expense.total_amount))
            ws.cell(row=row_num, column=13, value=expense.description)
            ws.cell(row=row_num, column=14, value=expense.approved_by.get_full_name() if expense.approved_by else '')
            ws.cell(row=row_num, column=15, value=expense.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Summary row
        if expenses.exists():
            summary_row = len(expenses) + 3
            ws.cell(row=summary_row, column=11, value="Total:").font = Font(bold=True)
            ws.cell(row=summary_row, column=12, value=sum(float(e.total_amount) for e in expenses)).font = Font(bold=True)
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'port_expenses_{start_date}_to_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

class DigitalVoucherExcelExportView(generics.GenericAPIView):
    """
    Export Digital Vouchers to Excel - Weekly Report for Accountants
    """
    permission_classes = [IsAccountantOrAdmin]
    
    def get(self, request):
        # Get week parameter or default to current week
        week_str = request.query_params.get('week')
        if week_str:
            try:
                start_date = datetime.strptime(week_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        else:
            # Default to current week (Monday to Sunday)
            today = timezone.now().date()
            start_date = today - timedelta(days=today.weekday())
        
        end_date = start_date + timedelta(days=6)
        
        # Get logged digital vouchers for the week
        vouchers = DigitalVoucher.objects.filter(
            status='logged',
            created_at__date__range=[start_date, end_date]
        ).order_by('created_at')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Digital Vouchers"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center")
        
        # Headers
        headers = [
            'Date', 'User', 'Expense Category', 'Amount', 'Remarks',
            'Approved By', 'Logged By', 'Tally Reference', 'Created At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, voucher in enumerate(vouchers, 2):
            ws.cell(row=row_num, column=1, value=voucher.date_time.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=2, value=voucher.user.get_full_name())
            ws.cell(row=row_num, column=3, value=voucher.get_expense_category_display())
            ws.cell(row=row_num, column=4, value=float(voucher.amount))
            ws.cell(row=row_num, column=5, value=voucher.remarks or '')
            ws.cell(row=row_num, column=6, value=voucher.approved_by.get_full_name() if voucher.approved_by else '')
            ws.cell(row=row_num, column=7, value=voucher.logged_by.get_full_name() if voucher.logged_by else '')
            ws.cell(row=row_num, column=8, value=voucher.tally_reference or '')
            ws.cell(row=row_num, column=9, value=voucher.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Summary row
        if vouchers.exists():
            summary_row = len(vouchers) + 3
            ws.cell(row=summary_row, column=3, value="Total:").font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=sum(float(v.amount) for v in vouchers)).font = Font(bold=True)
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'digital_vouchers_{start_date}_to_{end_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

class WalletHoldersView(generics.GenericAPIView):
    """
    View to get users who have wallets (for wallet top-up dropdown)
    """
    permission_classes = [CanManageWallets]
    
    def get(self, request):
        from authentication.models import User
        
        # Get all users except accountants
        wallet_holders = User.objects.filter(
            role__in=['admin', 'manager', 'supervisor'],
            is_active=True
        )
        
        # Add current balance for each holder
        holders_with_balance = []
        for user in wallet_holders:
            balance = Wallet.get_balance(user)
            holders_with_balance.append({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'role': user.role,
                'current_balance': str(balance)  # Convert to string for JSON serialization
            })
        
        return Response({
            'wallet_holders': holders_with_balance
        })

class ApprovalWorkflowView(generics.GenericAPIView):
    """
    Enhanced approval workflow view with better status tracking
    """
    permission_classes = [IsManagerOrAdmin]
    
    def get(self, request):
        user = request.user
        
        # Get pending items for approval
        pending_expenses = []
        pending_vouchers = []
        
        if user.is_manager or user.is_admin:
            # Port expenses pending manager/admin approval
            pending_expenses = PortExpense.objects.filter(
                status='submitted'
            ).select_related('user').values(
                'id', 'vehicle', 'vehicle_number', 'total_amount', 
                'date_time', 'user__username', 'created_at'
            )
            
            # Digital vouchers pending admin approval  
            if user.is_admin:
                pending_vouchers = DigitalVoucher.objects.filter(
                    status='submitted'
                ).select_related('user').values(
                    'id', 'expense_category', 'amount', 'date_time',
                    'user__username', 'created_at'
                )
        
        if user.is_accountant:
            # Port expenses pending accountant finalization
            pending_expenses = PortExpense.objects.filter(
                status='approved'
            ).select_related('user', 'reviewed_by').values(
                'id', 'vehicle', 'vehicle_number', 'total_amount',
                'date_time', 'user__username', 'reviewed_by__username', 'created_at'
            )
            
            # Digital vouchers pending accountant logging
            pending_vouchers = DigitalVoucher.objects.filter(
                status='approved'
            ).select_related('user', 'approved_by').values(
                'id', 'expense_category', 'amount', 'date_time',
                'user__username', 'approved_by__username', 'created_at'
            )
        
        return Response({
            'pending_expenses': list(pending_expenses),
            'pending_vouchers': list(pending_vouchers),
            'user_role': user.role,
            'approval_counts': {
                'expenses': len(pending_expenses),
                'vouchers': len(pending_vouchers),
                'total': len(pending_expenses) + len(pending_vouchers)
            }
        })

class BulkApprovalView(generics.GenericAPIView):
    """
    Bulk approval view for multiple expenses or vouchers
    """
    permission_classes = [IsManagerOrAdmin]
    
    def post(self, request):
        user = request.user
        item_type = request.data.get('type')  # 'expense' or 'voucher'
        item_ids = request.data.get('ids', [])
        action = request.data.get('action')  # 'approve', 'reject', 'finalize', 'log'
        comments = request.data.get('comments', '')
        
        if not item_type or not item_ids or not action:
            return Response(
                {'error': 'type, ids, and action are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        results = []
        
        if item_type == 'expense':
            expenses = PortExpense.objects.filter(id__in=item_ids)
            
            for expense in expenses:
                try:
                    if action == 'approve' and (user.is_manager or user.is_admin):
                        if expense.status == 'submitted':
                            expense.status = 'approved'
                            expense.reviewed_by = user
                            expense.review_comments = comments
                            expense.save()
                            results.append({'id': expense.id, 'status': 'approved'})
                        else:
                            results.append({'id': expense.id, 'error': 'Invalid status for approval'})
                    
                    elif action == 'finalize' and user.is_accountant:
                        if expense.status == 'approved':
                            expense.status = 'finalized'
                            expense.approved_by = user
                            expense.save()
                            results.append({'id': expense.id, 'status': 'finalized'})
                        else:
                            results.append({'id': expense.id, 'error': 'Must be approved before finalizing'})
                    
                    elif action == 'reject':
                        expense.status = 'rejected'
                        expense.reviewed_by = user
                        expense.review_comments = comments
                        expense.save()
                        results.append({'id': expense.id, 'status': 'rejected'})
                    
                    else:
                        results.append({'id': expense.id, 'error': 'Insufficient permissions or invalid action'})
                
                except Exception as e:
                    results.append({'id': expense.id, 'error': str(e)})
        
        elif item_type == 'voucher':
            vouchers = DigitalVoucher.objects.filter(id__in=item_ids)
            
            for voucher in vouchers:
                try:
                    if action == 'approve' and user.is_admin:
                        if voucher.status == 'submitted':
                            voucher.status = 'approved'
                            voucher.approved_by = user
                            voucher.approval_comments = comments
                            voucher.save()
                            results.append({'id': voucher.id, 'status': 'approved'})
                        else:
                            results.append({'id': voucher.id, 'error': 'Invalid status for approval'})
                    
                    elif action == 'log' and user.is_accountant:
                        if voucher.status == 'approved':
                            tally_ref = request.data.get('tally_reference', f'TALLY-{voucher.id}')
                            voucher.status = 'logged'
                            voucher.logged_by = user
                            voucher.tally_reference = tally_ref
                            voucher.save()
                            
                            # Create Tally log entry
                            TallyLog.objects.create(
                                entry_type='voucher',
                                reference_id=str(voucher.id),
                                tally_voucher_number=tally_ref,
                                amount=voucher.amount,
                                description=f"Digital voucher - {voucher.expense_category}",
                                logged_by=user
                            )
                            
                            results.append({'id': voucher.id, 'status': 'logged'})
                        else:
                            results.append({'id': voucher.id, 'error': 'Must be approved before logging'})
                    
                    elif action == 'decline':
                        voucher.status = 'declined'
                        voucher.approved_by = user
                        voucher.approval_comments = comments
                        voucher.save()
                        results.append({'id': voucher.id, 'status': 'declined'})
                    
                    else:
                        results.append({'id': voucher.id, 'error': 'Insufficient permissions or invalid action'})
                
                except Exception as e:
                    results.append({'id': voucher.id, 'error': str(e)})
        
        return Response({
            'results': results,
            'processed': len(results),
            'successful': len([r for r in results if 'error' not in r])
        })

class AllPortExpensesView(generics.ListAPIView):
    """
    List all port expenses for approval workflows
    """
    serializer_class = PortExpenseSerializer
    permission_classes = []  # Allow all authenticated users since we handle role-based filtering in get_queryset
    
    def get_queryset(self):
        user = self.request.user
        
        # Accountants see approved expenses ready for finalization
        if hasattr(user, 'role') and user.role == 'accountant':
            return PortExpense.objects.filter(status='approved').order_by('-created_at')
        
        # Managers and Admins see all expenses
        if hasattr(user, 'role') and user.role in ['manager', 'admin']:
            return PortExpense.objects.all().order_by('-created_at')
        
        # Supervisors see only their own expenses
        if hasattr(user, 'role') and user.role == 'supervisor':
            return PortExpense.objects.filter(user=user).order_by('-created_at')
        
        return PortExpense.objects.none()

class AllDigitalVouchersView(generics.ListAPIView):
    """
    List all digital vouchers for approval workflows
    """
    serializer_class = DigitalVoucherSerializer
    permission_classes = []  # Allow all authenticated users since we handle role-based filtering in get_queryset
    
    def get_queryset(self):
        user = self.request.user
        
        # Accountants see approved vouchers ready for Tally logging
        if hasattr(user, 'role') and user.role == 'accountant':
            return DigitalVoucher.objects.filter(status='approved').order_by('-created_at')
        
        # Admins see all vouchers
        if hasattr(user, 'role') and user.role == 'admin':
            return DigitalVoucher.objects.all().order_by('-created_at')
        
        # Managers and Supervisors see only their own vouchers
        if hasattr(user, 'role') and user.role in ['manager', 'supervisor']:
            return DigitalVoucher.objects.filter(user=user).order_by('-created_at')
        
        return DigitalVoucher.objects.none()
